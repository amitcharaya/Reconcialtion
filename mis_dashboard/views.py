"""
View/controller logic for the mis_dashboard application. Views receive HTTP requests, call service-layer code, and render templates or redirects.

Professional note:
    This project follows a simple separation of concerns:
    models store data, forms validate input, views control request/response flow,
    and services contain business rules such as import, reconciliation, dashboard
    summaries, dispute creation, and Excel generation.
"""

import json
from .services.upload_dashboard_service import UploadDashboardService
from django.http import HttpResponse
from django.shortcuts import render

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reconciliation.models import ATMReconciliationResult

from .services.dashboard_service import (
    MISDashboardService,
    MISDashboardFilterService,
)

from rgcs_reconciliation.models import RGCSReconciliationResult
from .services.rgcs_dashboard_service import RGCSDashboardService

def style_excel_sheet(sheet):
    header_fill = PatternFill(
        start_color="2C3E50",
        end_color="2C3E50",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column

        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        sheet.column_dimensions[
            get_column_letter(column)
        ].width = max_length + 3


def dashboard_view(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    summary = MISDashboardService.enterprise_summary(
        from_date,
        to_date
    )

    source_financial_summary = MISDashboardService.source_financial_summary(
        from_date,
        to_date
    )

    status_summary = MISDashboardService.status_summary(
        from_date,
        to_date
    )

    dispute_summary = MISDashboardService.dispute_status_summary(
        from_date,
        to_date
    )

    daily_trend = MISDashboardService.daily_reconciliation_trend(
        from_date,
        to_date
    )

    exception_summary = MISDashboardService.exception_summary(
        from_date,
        to_date
    )

    recent_disputes = MISDashboardService.recent_disputes(
        from_date,
        to_date
    )

    status_labels = [row["status"] for row in status_summary]
    status_values = [row["total"] for row in status_summary]

    trend_labels = [
        row["day"].strftime("%d-%m-%Y")
        if row["day"] else ""
        for row in daily_trend
    ]

    trend_total = [row["total"] for row in daily_trend]
    trend_matched = [row["matched"] for row in daily_trend]
    trend_unmatched = [row["unmatched"] for row in daily_trend]

    context = {
        "summary": summary,
        "source_financial_summary": source_financial_summary,

        "status_summary": status_summary,
        "dispute_summary": dispute_summary,
        "exception_summary": exception_summary,
        "recent_disputes": recent_disputes,

        "selected_from_date": from_date,
        "selected_to_date": to_date,

        "status_labels": json.dumps(status_labels),
        "status_values": json.dumps(status_values),

        "trend_labels": json.dumps(trend_labels),
        "trend_total": json.dumps(trend_total),
        "trend_matched": json.dumps(trend_matched),
        "trend_unmatched": json.dumps(trend_unmatched),
    }

    return render(
        request,
        "mis_dashboard/dashboard.html",
        context
    )


def reconciliation_report_view(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    status = request.GET.get("status")
    stan = request.GET.get("stan")
    rrn = request.GET.get("rrn")

    records = MISDashboardFilterService.filtered_reconciliation_queryset(
        from_date=from_date,
        to_date=to_date,
        status=status,
        stan=stan,
        rrn=rrn,
    ).order_by("-reconciled_at")

    context = {
        "records": records,
        "status_choices": ATMReconciliationResult.STATUS_CHOICES,

        "selected_from_date": from_date,
        "selected_to_date": to_date,
        "selected_status": status,
        "selected_stan": stan,
        "selected_rrn": rrn,
    }

    return render(
        request,
        "mis_dashboard/reconciliation_report.html",
        context
    )


def dispute_report_view(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    case_status = request.GET.get("case_status")
    stan = request.GET.get("stan")
    rrn = request.GET.get("rrn")

    disputes = MISDashboardFilterService.filtered_dispute_queryset(
        from_date=from_date,
        to_date=to_date,
        case_status=case_status,
        stan=stan,
        rrn=rrn,
    ).order_by("-created_at")

    case_status_choices = [
        "OPEN",
        "IN_PROGRESS",
        "CLOSED",
        "REJECTED",
    ]

    context = {
        "disputes": disputes,
        "case_status_choices": case_status_choices,

        "selected_from_date": from_date,
        "selected_to_date": to_date,
        "selected_case_status": case_status,
        "selected_stan": stan,
        "selected_rrn": rrn,
    }

    return render(
        request,
        "mis_dashboard/dispute_report.html",
        context
    )


def download_reconciliation_excel(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    status = request.GET.get("status")
    stan = request.GET.get("stan")
    rrn = request.GET.get("rrn")

    records = MISDashboardFilterService.filtered_reconciliation_queryset(
        from_date=from_date,
        to_date=to_date,
        status=status,
        stan=stan,
        rrn=rrn,
    ).order_by("-reconciled_at")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reconciliation Report"

    headers = [
        "Transaction Date",
        "STAN",
        "RRN",
        "CBS Present",
        "NDPG Present",
        "Switch Present",
        "CBS Amount",
        "NDPG Amount",
        "Switch Amount",
        "Matched By",
        "Status",
        "Remarks",
        "Reconciled At",
    ]

    sheet.append(headers)

    for row in records:
        transaction_date = ""

        if row.transaction_date:
            transaction_date = row.transaction_date.replace(
                tzinfo=None
            )

        sheet.append([
            transaction_date,
            row.stan_no,
            row.rrn,
            "Yes" if row.cbs_present else "No",
            "Yes" if row.ndpg_present else "No",
            "Yes" if row.switch_present else "No",
            row.cbs_amount,
            row.ndpg_amount,
            row.switch_amount,
            row.matched_by,
            row.status,
            row.remarks,
            row.reconciled_at.strftime("%d-%m-%Y %H:%M:%S")
            if row.reconciled_at else "",
        ])

    style_excel_sheet(sheet)

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="reconciliation_report.xlsx"'
    )

    workbook.save(response)

    return response


def download_dispute_excel(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    case_status = request.GET.get("case_status")
    stan = request.GET.get("stan")
    rrn = request.GET.get("rrn")

    disputes = MISDashboardFilterService.filtered_dispute_queryset(
        from_date=from_date,
        to_date=to_date,
        case_status=case_status,
        stan=stan,
        rrn=rrn,
    ).order_by("-created_at")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dispute Report"

    headers = [
        "Transaction Date",
        "STAN",
        "RRN",
        "Account No",
        "Disputed Amount",
        "Source Status",
        "Dispute Reason",
        "Case Status",
        "Assigned To",
        "Remarks",
        "Created At",
        "Updated At",
    ]

    sheet.append(headers)

    for row in disputes:
        sheet.append([
            row.transaction_date.strftime("%d-%m-%Y")
            if row.transaction_date else "",
            row.stan_no,
            row.rrn,
            row.account_no,
            row.disputed_amount,
            row.source_status,
            row.dispute_reason,
            row.case_status,
            row.assigned_to,
            row.remarks,
            row.created_at.strftime("%d-%m-%Y %H:%M:%S")
            if row.created_at else "",
            row.updated_at.strftime("%d-%m-%Y %H:%M:%S")
            if row.updated_at else "",
        ])

    style_excel_sheet(sheet)

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="dispute_report.xlsx"'
    )

    workbook.save(response)

    return response

def upload_monitoring_view(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    source = request.GET.get("source")
    status = request.GET.get("status")

    upload_records = UploadDashboardService.get_upload_records(
        from_date=from_date,
        to_date=to_date,
    )

    if source:
        upload_records = [
            row for row in upload_records
            if row["source"] == source
        ]

    if status:
        upload_records = [
            row for row in upload_records
            if row["upload_status"] == status
        ]

    context = {
        "upload_records": upload_records,
        "upload_summary": UploadDashboardService.upload_summary(from_date, to_date),
        "source_summary": UploadDashboardService.source_summary(from_date, to_date),

        "selected_from_date": from_date,
        "selected_to_date": to_date,
        "selected_source": source,
        "selected_status": status,
    }

    return render(
        request,
        "mis_dashboard/upload_monitoring.html",
        context
    )
# RGCS dashboard view. Date filters are also used to show disputes created
# within the selected range, giving operations teams a clear exposure view.
def rgcs_dashboard_view(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    context = {
        "summary": RGCSDashboardService.enterprise_summary(from_date, to_date),
        "status_summary": RGCSDashboardService.status_summary(from_date, to_date),
        "upload_summary": RGCSDashboardService.upload_summary(from_date, to_date),
        "dispute_status_summary": RGCSDashboardService.dispute_status_summary(from_date, to_date),
        "recent_disputes": RGCSDashboardService.recent_disputes(from_date, to_date),

        "selected_from_date": from_date,
        "selected_to_date": to_date,
    }

    return render(
        request,
        "mis_dashboard/rgcs_dashboard.html",
        context
    )



# Downloads the RGCS reconciliation report using the same filters selected
# on the report screen, so the Excel file always matches the visible report.
def download_rgcs_reconciliation_excel(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    status = request.GET.get("status")
    rrn = request.GET.get("rrn")

    records = RGCSDashboardService.filtered_reconciliation_queryset(
        from_date=from_date,
        to_date=to_date,
        status=status,
        rrn=rrn,
    ).order_by("-transaction_date", "rrn")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RGCS Reconciliation"

    headers = [
        "Transaction Date",
        "RRN",
        "CBS Transaction ID",
        "Switch Transaction ID",
        "NDPG Transaction ID",
        "CBS Amount",
        "Switch Amount",
        "NDPG Amount",
        "Status",
        "Remarks",
        "Reconciled At",
    ]

    sheet.append(headers)

    for row in records:
        sheet.append([
            row.transaction_date.strftime("%d-%m-%Y")
            if row.transaction_date else "",
            row.rrn,
            row.cbs_transaction_id or "",
            row.switch_transaction_id or "",
            row.ndpg_transaction_id or "",
            row.cbs_amount,
            row.switch_amount,
            row.ndpg_amount,
            row.status,
            row.remarks or "",
            row.reconciled_at.strftime("%d-%m-%Y %H:%M:%S")
            if row.reconciled_at else "",
        ])

    style_excel_sheet(sheet)

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    filename_from = from_date or "all"
    filename_to = to_date or "all"
    response["Content-Disposition"] = (
        f'attachment; filename="rgcs_reconciliation_report_{filename_from}_to_{filename_to}.xlsx"'
    )

    workbook.save(response)

    return response

def rgcs_reconciliation_report_view(request):
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    status = request.GET.get("status")
    rrn = request.GET.get("rrn")

    records = RGCSDashboardService.filtered_reconciliation_queryset(
        from_date=from_date,
        to_date=to_date,
        status=status,
        rrn=rrn,
    ).order_by("-transaction_date", "rrn")

    context = {
        "records": records,
        "status_choices": RGCSReconciliationResult.STATUS_CHOICES,

        "selected_from_date": from_date,
        "selected_to_date": to_date,
        "selected_status": status,
        "selected_rrn": rrn,
    }

    return render(
        request,
        "mis_dashboard/rgcs_reconciliation_report.html",
        context
    )
from datetime import date
from .services.upload_workflow_service import UploadWorkflowService


def home_page(request):

   selected_date = request.GET.get("transaction_date") or request.GET.get("date") or date.today().isoformat()

   workflows = UploadWorkflowService.build(selected_date)

   return render(
        request,
        "mis_dashboard/home.html",
        {
            "selected_date": selected_date,
            "workflows": workflows,
        }
    )
